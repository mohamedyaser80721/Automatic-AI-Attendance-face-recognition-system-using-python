import cv2
import face_recognition
import pickle
import numpy as np
from datetime import datetime
import requests
from openpyxl import Workbook, load_workbook
import os
from supabase import create_client, Client
import io

# ---------- Face Recognition Setup ----------
ENCODINGS_PATH = "encodings.pkl"
TOLERANCE = 0.45   # lower = stricter matching

# load known encodings
with open(ENCODINGS_PATH, "rb") as f:
    data = pickle.load(f)
known_encodings = data["encodings"]
known_names = data["names"]

# ---------- OpenCage Setup ----------
OPENCAGE_API_KEY = "--YOUR--API---"

def get_location():
    try:
        ipinfo = requests.get("https://ipinfo.io/json").json()
        lat, lng = ipinfo["loc"].split(",")
        url = f"https://api.opencagedata.com/geocode/v1/json?q={lat}+{lng}&key={OPENCAGE_API_KEY}"
        res = requests.get(url).json()
        if res["results"]:
            return res["results"][0]["formatted"]
    except:
        pass
    return "Unknown Location"

# ---------- Excel Setup ----------
EXCEL_FILE = "attendance.xlsx"

if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"
    ws.append(["Name", "Timestamp", "Location", "Image URL"])  # added image URL
    wb.save(EXCEL_FILE)
else:
    print("file already exist..now opening camera")

def save_to_excel(name, timestamp, location, image_url):
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Attendance"]
    ws.append([name, timestamp, location, image_url])
    wb.save(EXCEL_FILE)

# ---------- Supabase Setup ----------
SUPABASE_URL = "---YOUR SUPABASE URL---"
SUPABASE_KEY = "---YOUR SUPABASE KEY---"
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

def save_to_supabase(name, timestamp, location, face_img):
    # Convert face image to JPG bytes
    _, buffer = cv2.imencode(".jpg", face_img)
    img_bytes = io.BytesIO(buffer)

    # Create unique filename
    filename = f"{name}_{timestamp.replace(':','-').replace(' ','_')}.jpg"

    # Upload to Supabase storage bucket
    supabase.storage.from_("attendance-photos").upload(filename, img_bytes.getvalue())

    # Get public URL of image
    image_url = supabase.storage.from_("attendance-photos").get_public_url(filename)

    # Insert into Supabase table
    supabase.table("attendance").insert({
        "name": name,
        "timestamp": timestamp,
        "location": location,
        "image_url": image_url
    }).execute()

    print("📤 Supabase saved:", name, timestamp, location, image_url)
    return image_url

# ---------- Webcam Setup ----------
video = cv2.VideoCapture(0)
if not video.isOpened():
    raise RuntimeError("Could not open webcam")

print("Starting webcam. Press 'q' to save & quit.")

while True:
    ret, frame = video.read()
    if not ret:
        break

    # resize for speed
    small = cv2.resize(frame, (0, 0), fx=0.5, fy=0.5)
    rgb_small = cv2.cvtColor(small, cv2.COLOR_BGR2RGB)

    # detect and encode faces
    face_locations = face_recognition.face_locations(rgb_small, model="hog")
    face_encodings = face_recognition.face_encodings(rgb_small, face_locations)

    detected_faces = []  # store multiple faces

    for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
        name = "Unknown"
        distances = face_recognition.face_distance(known_encodings, face_encoding)
        if len(distances) > 0:
            best_idx = np.argmin(distances)
            if distances[best_idx] <= TOLERANCE:
                name = known_names[best_idx]

        # scale coords back
        top *= 2; right *= 2; bottom *= 2; left *= 2
        cv2.rectangle(frame, (left, top), (right, bottom), (0, 255, 0), 2)

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        cv2.putText(frame, f"{name} | {timestamp}", (left, bottom + 25),
                    cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)

        # crop the face
        face_img = frame[top:bottom, left:right]

        # store detected face info with image
        detected_faces.append({"name": name, "timestamp": timestamp, "face_img": face_img})

    cv2.imshow("Face Recognition", frame)

    key = cv2.waitKey(1) & 0xFF
    if key == ord("q"):
        location = get_location()

        for face in detected_faces:
            # Save cropped face to Supabase
            image_url = save_to_supabase(face["name"], face["timestamp"], location, face["face_img"])

            # Save metadata to Excel
            save_to_excel(face["name"], face["timestamp"], location, image_url)

            print("✅ Data saved:", face["name"], face["timestamp"], location)

        break

video.release()
cv2.destroyAllWindows()
